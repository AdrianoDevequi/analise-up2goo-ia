import os
import threading
import string
import random
from datetime import datetime
from contextlib import contextmanager
from functools import wraps

import psycopg2
import psycopg2.extras
from io import BytesIO
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, jsonify, g, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

from crawler import crawl_website
from analyzer import analyze_page, get_ai_suggestions
from sf_importer import process_sf_csv, crawl_with_sf_cli, sf_cli_available
from sitemap_parser import get_sitemap_urls, guess_sitemap_url

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(32).hex())


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def get_db_config():
    # Railway (e outros PaaS) injetam DATABASE_URL automaticamente
    database_url = os.environ.get('DATABASE_URL', '')
    if database_url:
        from urllib.parse import urlparse
        # psycopg2 exige "postgresql://", Railway às vezes envia "postgres://"
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        p = urlparse(database_url)
        return {
            'host': p.hostname,
            'port': p.port or 5432,
            'dbname': p.path.lstrip('/'),
            'user': p.username,
            'password': p.password,
        }
    return {
        'host': os.environ.get('DB_HOST', 'localhost'),
        'port': int(os.environ.get('DB_PORT', 5432)),
        'dbname': os.environ.get('DB_NAME', 'analise_ia'),
        'user': os.environ.get('DB_USER', 'analise_user'),
        'password': os.environ.get('DB_PASSWORD', 'analise_pass123'),
    }


def get_db():
    """Returns a database connection attached to Flask's g context."""
    if '_db' not in g:
        g._db = psycopg2.connect(
            **get_db_config(),
            cursor_factory=psycopg2.extras.RealDictCursor
        )
    return g._db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('_db', None)
    if db is not None:
        db.close()


@contextmanager
def get_thread_db():
    """Database connection for background threads (no Flask context)."""
    conn = psycopg2.connect(
        **get_db_config(),
        cursor_factory=psycopg2.extras.RealDictCursor
    )
    try:
        yield conn
    finally:
        conn.close()


def init_db():
    """Create tables and default admin user. Retries until DB is reachable."""
    import time

    # Warn early if DATABASE_URL is missing in a non-local environment
    if not os.environ.get('DATABASE_URL') and not os.environ.get('DB_HOST'):
        print('[INIT] AVISO: DATABASE_URL não definida. '
              'No Railway, adicione o plugin PostgreSQL ao projeto.')

    max_attempts = 15
    for attempt in range(1, max_attempts + 1):
        try:
            conn_test = psycopg2.connect(**get_db_config())
            conn_test.close()
            break
        except psycopg2.OperationalError as e:
            if attempt == max_attempts:
                print(f'[INIT] Banco inacessível após {max_attempts} tentativas. Abortando.')
                raise
            wait = min(attempt * 2, 15)
            print(f'[INIT] Banco não disponível (tentativa {attempt}/{max_attempts}). '
                  f'Aguardando {wait}s... ({e})'.split('\n')[0])
            time.sleep(wait)

    print('[INIT] Banco de dados conectado.')
    with get_thread_db() as conn:
        with conn.cursor() as cur:
            cur.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'client',
                    created_at TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS projects (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    url TEXT NOT NULL,
                    sitemap_url TEXT,
                    client_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    plain_password TEXT,
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS analyses (
                    id SERIAL PRIMARY KEY,
                    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                    status TEXT DEFAULT 'pending',
                    total_pages INTEGER DEFAULT 0,
                    pages_expected INTEGER DEFAULT 0,
                    total_issues INTEGER DEFAULT 0,
                    high_issues INTEGER DEFAULT 0,
                    medium_issues INTEGER DEFAULT 0,
                    low_issues INTEGER DEFAULT 0,
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP,
                    error_message TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS pages (
                    id SERIAL PRIMARY KEY,
                    analysis_id INTEGER NOT NULL REFERENCES analyses(id) ON DELETE CASCADE,
                    url TEXT NOT NULL,
                    title TEXT DEFAULT '',
                    status_code INTEGER DEFAULT 0,
                    word_count INTEGER DEFAULT 0,
                    load_time REAL DEFAULT 0,
                    issue_count INTEGER DEFAULT 0,
                    analyzed_at TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS issues (
                    id SERIAL PRIMARY KEY,
                    page_id INTEGER NOT NULL REFERENCES pages(id) ON DELETE CASCADE,
                    category TEXT NOT NULL,
                    severity TEXT NOT NULL,
                    title TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    current_value TEXT DEFAULT '',
                    suggestion TEXT DEFAULT '',
                    ai_suggestion JSONB
                );
            ''')
            conn.commit()

            # Migrations for existing deployments
            cur.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='projects' AND column_name='sitemap_url'
                    ) THEN
                        ALTER TABLE projects ADD COLUMN sitemap_url TEXT;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='projects' AND column_name='use_playwright'
                    ) THEN
                        ALTER TABLE projects ADD COLUMN use_playwright BOOLEAN DEFAULT FALSE;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='analyses' AND column_name='pages_expected'
                    ) THEN
                        ALTER TABLE analyses ADD COLUMN pages_expected INTEGER DEFAULT 0;
                    END IF;
                END $$;
            """)
            conn.commit()

            # Default admin
            admin_email = os.environ.get('ADMIN_EMAIL', 'admin@seudominio.com')
            admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
            admin_name = os.environ.get('ADMIN_NAME', 'Administrador')

            cur.execute('SELECT id FROM users WHERE role = %s LIMIT 1', ('admin',))
            if not cur.fetchone():
                cur.execute(
                    'INSERT INTO users (name, email, password_hash, role) VALUES (%s, %s, %s, %s)',
                    (admin_name, admin_email, generate_password_hash(admin_password), 'admin')
                )
                conn.commit()
                print(f'[INIT] Admin criado: {admin_email} / {admin_password}')

            # Mark any stale running/pending analyses as stopped (server restarted)
            cur.execute(
                """UPDATE analyses SET status='stopped', completed_at=%s,
                   error_message='Interrompido por reinicialização do servidor'
                   WHERE status IN ('running', 'pending')""",
                (datetime.now(),)
            )
            affected = cur.rowcount
            conn.commit()
            if affected:
                print(f'[INIT] {affected} análise(s) interrompida(s) por reinicialização.')


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Analysis cancellation registry
# ---------------------------------------------------------------------------

_cancel_set: set = set()   # analysis IDs requested to stop


# Auth helpers
# ---------------------------------------------------------------------------

def generate_password(length=10):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=length))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Acesso restrito a administradores.', 'danger')
            return redirect(url_for('client_dashboard'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Background analysis
# ---------------------------------------------------------------------------

def run_analysis_background(analysis_id, project_url, project_id=None, sitemap_url=None, use_playwright=False):
    """Runs the full crawl + SEO analysis in a background thread."""
    from crawler import fetch_page, fetch_page_playwright
    with get_thread_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    'UPDATE analyses SET status=%s, started_at=%s WHERE id=%s',
                    ('running', datetime.now(), analysis_id)
                )
                conn.commit()

                fetch_fn = fetch_page_playwright if use_playwright else fetch_page

                total_issues = 0
                high_issues = 0
                medium_issues = 0
                low_issues = 0
                pages_done = 0
                has_api_key = bool(os.environ.get('GEMINI_API_KEY', '').strip())
                import json as _json

                def _process_page(page_data):
                    """Analyze, save, and update progress for one page. Returns False if cancelled."""
                    nonlocal total_issues, high_issues, medium_issues, low_issues, pages_done

                    issues, word_count = analyze_page(page_data)

                    ai_data = None
                    if has_api_key and pages_done < 10:
                        ai_data = get_ai_suggestions(page_data)

                    cur.execute(
                        '''INSERT INTO pages (analysis_id, url, title, status_code, word_count, load_time, issue_count)
                           VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id''',
                        (analysis_id, page_data['url'], page_data.get('title', ''),
                         page_data.get('status_code', 0), word_count,
                         page_data.get('load_time', 0), len(issues))
                    )
                    page_id = cur.fetchone()['id']

                    for issue in issues:
                        ai_json = _json.dumps(ai_data, ensure_ascii=False) if ai_data else None
                        cur.execute(
                            '''INSERT INTO issues (page_id, category, severity, title,
                                                   description, current_value, suggestion, ai_suggestion)
                               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                            (page_id, issue['category'], issue['severity'], issue['title'],
                             issue.get('description', ''), issue.get('current_value', ''),
                             issue.get('suggestion', ''), ai_json)
                        )
                        total_issues += 1
                        if issue['severity'] == 'high':     high_issues += 1
                        elif issue['severity'] == 'medium': medium_issues += 1
                        else:                               low_issues += 1

                    pages_done += 1
                    cur.execute('UPDATE analyses SET total_pages=%s, total_issues=%s WHERE id=%s',
                                (pages_done, total_issues, analysis_id))
                    conn.commit()

                    if analysis_id in _cancel_set:
                        _cancel_set.discard(analysis_id)
                        print(f'[ANALYSIS] Interrompida após {pages_done} páginas.')
                        # Only update counts — status was already set to 'stopped' by the stop route
                        cur.execute(
                            '''UPDATE analyses SET total_pages=%s, total_issues=%s,
                                   high_issues=%s, medium_issues=%s, low_issues=%s
                                   WHERE id=%s''',
                            (pages_done, total_issues, high_issues, medium_issues, low_issues, analysis_id)
                        )
                        conn.commit()
                        return False
                    return True

                # ── Modo sitemap: busca + analisa cada URL imediatamente ──
                if sitemap_url:
                    print(f'[ANALYSIS] Usando sitemap: {sitemap_url}')
                    # Signal UI that sitemap is loading (-1 = loading)
                    cur.execute('UPDATE analyses SET pages_expected=%s WHERE id=%s',
                                (-1, analysis_id))
                    conn.commit()
                    try:
                        sitemap_urls = get_sitemap_urls(sitemap_url, max_urls=500)
                        print(f'[ANALYSIS] {len(sitemap_urls)} URLs no sitemap')
                    except Exception as e:
                        print(f'[ANALYSIS] Falha no sitemap, crawl normal: {e}')
                        sitemap_urls = None

                    if sitemap_urls:
                        url_list = sitemap_urls
                        cur.execute('UPDATE analyses SET pages_expected=%s WHERE id=%s',
                                    (len(url_list), analysis_id))
                        conn.commit()

                        for url in url_list:
                            # Skip if already analyzed in the last hour
                            if project_id:
                                cur.execute('''
                                    SELECT pg.id, pg.title, pg.status_code, pg.word_count,
                                           pg.load_time, pg.issue_count
                                    FROM pages pg
                                    JOIN analyses a ON a.id = pg.analysis_id
                                    WHERE pg.url = %s
                                      AND a.project_id = %s
                                      AND a.id != %s
                                      AND pg.analyzed_at > NOW() - INTERVAL '1 hour'
                                    ORDER BY pg.analyzed_at DESC
                                    LIMIT 1
                                ''', (url, project_id, analysis_id))
                                cached = cur.fetchone()
                                if cached:
                                    print(f'[ANALYSIS] Cache hit (< 1h): {url}')
                                    # Copy page record
                                    cur.execute(
                                        '''INSERT INTO pages (analysis_id, url, title, status_code,
                                               word_count, load_time, issue_count)
                                           VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id''',
                                        (analysis_id, url, cached['title'], cached['status_code'],
                                         cached['word_count'], cached['load_time'], cached['issue_count'])
                                    )
                                    new_page_id = cur.fetchone()['id']
                                    # Copy issues
                                    cur.execute('SELECT * FROM issues WHERE page_id = %s', (cached['id'],))
                                    for iss in cur.fetchall():
                                        cur.execute(
                                            '''INSERT INTO issues (page_id, category, severity, title,
                                                   description, current_value, suggestion, ai_suggestion)
                                               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                                            (new_page_id, iss['category'], iss['severity'], iss['title'],
                                             iss['description'], iss['current_value'],
                                             iss['suggestion'], iss['ai_suggestion'])
                                        )
                                        total_issues += 1
                                        if iss['severity'] == 'high':     high_issues += 1
                                        elif iss['severity'] == 'medium': medium_issues += 1
                                        else:                              low_issues += 1
                                    pages_done += 1
                                    cur.execute('UPDATE analyses SET total_pages=%s, total_issues=%s WHERE id=%s',
                                                (pages_done, total_issues, analysis_id))
                                    conn.commit()
                                    if analysis_id in _cancel_set:
                                        _cancel_set.discard(analysis_id)
                                        return
                                    continue

                            try:
                                page_data = fetch_fn(url)
                                if not page_data:
                                    continue
                            except Exception as e:
                                print(f'[ANALYSIS] Erro ao buscar {url}: {e}')
                                continue
                            if not _process_page(page_data):
                                return  # cancelled
                        # final update and return early
                        cur.execute(
                            '''UPDATE analyses SET status=%s, completed_at=%s,
                                   total_pages=%s, total_issues=%s,
                                   high_issues=%s, medium_issues=%s, low_issues=%s WHERE id=%s''',
                            ('completed', datetime.now(), pages_done, total_issues,
                             high_issues, medium_issues, low_issues, analysis_id)
                        )
                        conn.commit()
                        return

                    # Sitemap falhou — fallback para crawl
                    pages = crawl_website(project_url, max_pages=50)
                else:
                    pages = crawl_website(project_url, max_pages=30)

                # ── Modo crawl: processa lista retornada pelo crawler ──
                cur.execute('UPDATE analyses SET pages_expected=%s WHERE id=%s',
                            (len(pages), analysis_id))
                conn.commit()

                for page_data in pages:
                    if not _process_page(page_data):
                        return  # cancelled

                cur.execute(
                    '''UPDATE analyses SET
                        status=%s, completed_at=%s,
                        total_pages=%s, total_issues=%s,
                        high_issues=%s, medium_issues=%s, low_issues=%s
                       WHERE id=%s''',
                    (
                        'completed', datetime.now(),
                        pages_done, total_issues,
                        high_issues, medium_issues, low_issues,
                        analysis_id
                    )
                )
                conn.commit()

            except Exception as e:
                print(f'[ANALYSIS ERROR] {e}')
                try:
                    cur.execute(
                        "UPDATE analyses SET status='error', error_message=%s WHERE id=%s",
                        (str(e), analysis_id)
                    )
                    conn.commit()
                except Exception:
                    pass


# ---------------------------------------------------------------------------
# Shared background helper: save SF/CLI results to DB
# ---------------------------------------------------------------------------

def _save_sf_results_to_db(conn, analysis_id, sf_results, has_api_key):
    """Persist Screaming Frog parsed results (pages + issues) to the database."""
    import json as _json
    total_issues = high_issues = medium_issues = low_issues = 0

    with conn.cursor() as cur:
        for page_data, issues in sf_results:
            ai_data = None
            if has_api_key and issues:
                # Build a lightweight page_data dict for AI suggestions
                ai_page = {
                    'url': page_data['url'],
                    'soup': None,
                    '_sf_title': page_data.get('title', ''),
                    '_sf_meta': page_data.get('_meta', ''),
                    '_sf_h1': page_data.get('_h1', ''),
                }
                ai_data = _get_ai_from_sf_data(ai_page)

            cur.execute(
                '''INSERT INTO pages (analysis_id, url, title, status_code, word_count, load_time, issue_count)
                   VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id''',
                (
                    analysis_id,
                    page_data['url'],
                    page_data.get('title', ''),
                    page_data.get('status_code', 0),
                    page_data.get('word_count', 0),
                    page_data.get('load_time', 0),
                    len(issues)
                )
            )
            page_id = cur.fetchone()['id']
            conn.commit()

            for issue in issues:
                ai_json = _json.dumps(ai_data, ensure_ascii=False) if ai_data else None
                cur.execute(
                    '''INSERT INTO issues (page_id, category, severity, title,
                                          description, current_value, suggestion, ai_suggestion)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                    (
                        page_id,
                        issue['category'], issue['severity'], issue['title'],
                        issue.get('description', ''), issue.get('current_value', ''),
                        issue.get('suggestion', ''), ai_json
                    )
                )
                total_issues += 1
                if issue['severity'] == 'high':   high_issues += 1
                elif issue['severity'] == 'medium': medium_issues += 1
                else:                               low_issues += 1

            conn.commit()

    return total_issues, high_issues, medium_issues, low_issues


def _get_ai_from_sf_data(page_data):
    """Call Gemini AI using SF metadata (title, meta, H1) instead of full HTML."""
    import os, re, json as _json
    api_key = os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        return None
    try:
        import google.generativeai as genai
        title = page_data.get('_sf_title') or page_data.get('title', '')
        meta  = page_data.get('_sf_meta', '')
        h1    = page_data.get('_sf_h1', '')
        url   = page_data.get('url', '')

        prompt = f"""Você é especialista em SEO e copywriting para o mercado brasileiro.
Analise os elementos desta página e sugira melhorias otimizadas para SEO.

URL: {url}
Título atual: {title or '(ausente)'}
Meta description atual: {meta or '(ausente)'}
H1 atual: {h1 or '(ausente)'}

Retorne APENAS JSON válido (sem markdown):
{{
  "titulo": "Título otimizado com 50-60 caracteres",
  "meta_description": "Meta description persuasiva com 150-160 caracteres e call-to-action",
  "h1": "H1 principal otimizado",
  "dica_conteudo": "Sugestão prática para melhorar o conteúdo desta página",
  "palavras_chave": ["kw1", "kw2", "kw3"]
}}

Use português brasileiro. Seja específico ao nicho identificado na URL/título."""

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        response = model.generate_content(prompt)
        text = response.text.strip()
        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            return _json.loads(m.group())
    except Exception as e:
        print(f'[AI/SF] {e}')
    return None


def run_sf_import_background(analysis_id, sf_results):
    """Background thread: persist pre-parsed SF results to DB."""
    with get_thread_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    'UPDATE analyses SET status=%s, started_at=%s WHERE id=%s',
                    ('running', datetime.now(), analysis_id)
                )
                conn.commit()
            except Exception as e:
                print(f'[SF IMPORT] Error setting running: {e}')
                return

        has_api_key = bool(os.environ.get('GEMINI_API_KEY', '').strip())
        total, high, medium, low = _save_sf_results_to_db(conn, analysis_id, sf_results, has_api_key)

        with conn.cursor() as cur:
            try:
                cur.execute(
                    '''UPDATE analyses SET status=%s, completed_at=%s,
                           total_pages=%s, total_issues=%s,
                           high_issues=%s, medium_issues=%s, low_issues=%s
                       WHERE id=%s''',
                    ('completed', datetime.now(),
                     len(sf_results), total, high, medium, low,
                     analysis_id)
                )
                conn.commit()
            except Exception as e:
                print(f'[SF IMPORT] Error completing: {e}')
                try:
                    cur.execute(
                        "UPDATE analyses SET status='error', error_message=%s WHERE id=%s",
                        (str(e), analysis_id)
                    )
                    conn.commit()
                except Exception:
                    pass


def run_sf_cli_background(analysis_id, project_url):
    """Background thread: run SF CLI, parse output, save to DB."""
    with get_thread_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    'UPDATE analyses SET status=%s, started_at=%s WHERE id=%s',
                    ('running', datetime.now(), analysis_id)
                )
                conn.commit()
            except Exception:
                return

        try:
            sf_results = crawl_with_sf_cli(project_url)
        except Exception as e:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE analyses SET status='error', error_message=%s WHERE id=%s",
                    (str(e), analysis_id)
                )
                conn.commit()
            return

        has_api_key = bool(os.environ.get('GEMINI_API_KEY', '').strip())
        total, high, medium, low = _save_sf_results_to_db(conn, analysis_id, sf_results, has_api_key)

        with conn.cursor() as cur:
            try:
                cur.execute(
                    '''UPDATE analyses SET status=%s, completed_at=%s,
                           total_pages=%s, total_issues=%s,
                           high_issues=%s, medium_issues=%s, low_issues=%s
                       WHERE id=%s''',
                    ('completed', datetime.now(),
                     len(sf_results), total, high, medium, low,
                     analysis_id)
                )
                conn.commit()
            except Exception as e:
                with conn.cursor() as cur2:
                    cur2.execute(
                        "UPDATE analyses SET status='error', error_message=%s WHERE id=%s",
                        (str(e), analysis_id)
                    )
                    conn.commit()


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('client_dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')

        db = get_db()
        with db.cursor() as cur:
            cur.execute('SELECT * FROM users WHERE email = %s', (email,))
            user = cur.fetchone()

        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['name'] = user['name']
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('client_dashboard'))

        flash('E-mail ou senha incorretos.', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------

@app.route('/admin')
@admin_required
def admin_dashboard():
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT COUNT(*) as total FROM projects')
        total_projects = cur.fetchone()['total']

        cur.execute("SELECT COUNT(*) as total FROM users WHERE role='client'")
        total_clients = cur.fetchone()['total']

        cur.execute("SELECT COUNT(*) as total FROM analyses WHERE status='completed'")
        total_analyses = cur.fetchone()['total']

        cur.execute('''
            SELECT p.id, p.name, p.url, p.created_at,
                   u.name as client_name, u.email as client_email,
                   (SELECT status FROM analyses WHERE project_id=p.id ORDER BY created_at DESC LIMIT 1) as last_status,
                   (SELECT completed_at FROM analyses WHERE project_id=p.id ORDER BY created_at DESC LIMIT 1) as last_analysis
            FROM projects p
            JOIN users u ON u.id = p.client_id
            ORDER BY p.created_at DESC
            LIMIT 10
        ''')
        recent_projects = cur.fetchall()

    return render_template('admin/dashboard.html',
                           total_projects=total_projects,
                           total_clients=total_clients,
                           total_analyses=total_analyses,
                           recent_projects=recent_projects)


@app.route('/admin/projetos')
@admin_required
def admin_projects():
    db = get_db()
    with db.cursor() as cur:
        cur.execute('''
            SELECT p.id, p.name, p.url, p.created_at,
                   u.name as client_name, u.email as client_email,
                   (SELECT status FROM analyses WHERE project_id=p.id ORDER BY created_at DESC LIMIT 1) as last_status,
                   (SELECT total_issues FROM analyses WHERE project_id=p.id AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1) as last_issues,
                   (SELECT completed_at FROM analyses WHERE project_id=p.id AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1) as last_analysis
            FROM projects p
            JOIN users u ON u.id = p.client_id
            ORDER BY p.created_at DESC
        ''')
        projects = cur.fetchall()

    return render_template('admin/projects.html', projects=projects)


@app.route('/admin/projetos/novo', methods=['GET', 'POST'])
@admin_required
def admin_new_project():
    if request.method == 'POST':
        project_name = request.form.get('project_name', '').strip()
        client_name = request.form.get('client_name', '').strip()
        client_email = request.form.get('client_email', '').strip().lower()
        project_url = request.form.get('project_url', '').strip()
        sitemap_url = request.form.get('sitemap_url', '').strip() or None
        use_playwright = request.form.get('use_playwright') == 'on'
        notes = request.form.get('notes', '').strip()

        if not all([project_name, client_name, client_email, project_url]):
            flash('Preencha todos os campos obrigatórios.', 'danger')
            return render_template('admin/new_project.html')

        # Normalize URL
        if not project_url.startswith(('http://', 'https://')):
            project_url = 'https://' + project_url

        # Auto-fill sitemap if not provided
        if not sitemap_url:
            sitemap_url = guess_sitemap_url(project_url)

        db = get_db()
        with db.cursor() as cur:
            # Check if client user exists
            cur.execute('SELECT id FROM users WHERE email = %s', (client_email,))
            existing_user = cur.fetchone()

            plain_password = generate_password(10)

            if existing_user:
                client_id = existing_user['id']
                # Update password so admin has the current one
                cur.execute(
                    'UPDATE users SET password_hash=%s, name=%s WHERE id=%s',
                    (generate_password_hash(plain_password), client_name, client_id)
                )
            else:
                cur.execute(
                    'INSERT INTO users (name, email, password_hash, role) VALUES (%s, %s, %s, %s) RETURNING id',
                    (client_name, client_email, generate_password_hash(plain_password), 'client')
                )
                client_id = cur.fetchone()['id']

            cur.execute(
                'INSERT INTO projects (name, url, sitemap_url, use_playwright, client_id, plain_password, notes) VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id',
                (project_name, project_url, sitemap_url, use_playwright, client_id, plain_password, notes)
            )
            project_id = cur.fetchone()['id']
            db.commit()

        flash(f'Projeto criado! Credenciais do cliente: {client_email} / {plain_password}', 'success')
        return redirect(url_for('admin_project_detail', project_id=project_id))

    return render_template('admin/new_project.html')


@app.route('/admin/projetos/<int:project_id>')
@admin_required
def admin_project_detail(project_id):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('''
            SELECT p.*, u.name as client_name, u.email as client_email
            FROM projects p
            JOIN users u ON u.id = p.client_id
            WHERE p.id = %s
        ''', (project_id,))
        project = cur.fetchone()

        if not project:
            flash('Projeto não encontrado.', 'danger')
            return redirect(url_for('admin_projects'))

        cur.execute('''
            SELECT * FROM analyses WHERE project_id = %s ORDER BY created_at DESC
        ''', (project_id,))
        analyses = cur.fetchall()

        latest_analysis = analyses[0] if analyses else None
        pages_data = []

        # Find the most recent completed/stopped analysis for pages display
        display_analysis = next(
            (a for a in analyses if a['status'] in ('completed', 'stopped')), None
        )

        if display_analysis:
            cur.execute('''
                SELECT pg.*,
                       COUNT(i.id) as total_issue_count,
                       COUNT(CASE WHEN i.severity='high' THEN 1 END) as high_count,
                       COUNT(CASE WHEN i.severity='medium' THEN 1 END) as medium_count,
                       COUNT(CASE WHEN i.severity='low' THEN 1 END) as low_count
                FROM pages pg
                LEFT JOIN issues i ON i.page_id = pg.id
                WHERE pg.analysis_id = %s
                GROUP BY pg.id
                ORDER BY COUNT(i.id) DESC
            ''', (display_analysis['id'],))
            pages_data = cur.fetchall()

    return render_template('admin/project_detail.html',
                           project=project,
                           analyses=analyses,
                           latest_analysis=latest_analysis,
                           display_analysis=display_analysis,
                           pages_data=pages_data)


@app.route('/admin/projetos/<int:project_id>/atualizar-sitemap', methods=['POST'])
@admin_required
def admin_update_sitemap(project_id):
    """Update crawler settings (sitemap_url + use_playwright) for an existing project."""
    sitemap_url = request.form.get('sitemap_url', '').strip() or None
    use_playwright = request.form.get('use_playwright') == 'on'
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            'UPDATE projects SET sitemap_url=%s, use_playwright=%s WHERE id=%s',
            (sitemap_url, use_playwright, project_id)
        )
        db.commit()
    flash('Configurações do rastreio salvas.', 'success')
    return redirect(url_for('admin_project_detail', project_id=project_id))


@app.route('/admin/projetos/<int:project_id>/analisar', methods=['POST'])
@admin_required
def admin_run_analysis(project_id):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT * FROM projects WHERE id = %s', (project_id,))
        project = cur.fetchone()

        if not project:
            return jsonify({'error': 'Projeto não encontrado'}), 404

        # Check if there's already a running analysis
        cur.execute(
            "SELECT id FROM analyses WHERE project_id=%s AND status IN ('pending','running')",
            (project_id,)
        )
        running = cur.fetchone()
        if running:
            return jsonify({'error': 'Já existe uma análise em andamento para este projeto'}), 400

        cur.execute(
            'INSERT INTO analyses (project_id, status) VALUES (%s, %s) RETURNING id',
            (project_id, 'pending')
        )
        analysis_id = cur.fetchone()['id']
        db.commit()

    # Start background thread
    thread = threading.Thread(
        target=run_analysis_background,
        args=(analysis_id, project['url']),
        kwargs={
            'project_id': project_id,
            'sitemap_url': project.get('sitemap_url') or None,
            'use_playwright': bool(project.get('use_playwright')),
        },
        daemon=True
    )
    thread.start()

    return jsonify({'analysis_id': analysis_id, 'status': 'started'})


@app.route('/admin/projetos/<int:project_id>/parar-analise', methods=['POST'])
@admin_required
def admin_stop_analysis(project_id):
    """Request cancellation of the running analysis for a project."""
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            "SELECT id, total_pages, total_issues, high_issues, medium_issues, low_issues FROM analyses WHERE project_id=%s AND status IN ('pending','running') ORDER BY created_at DESC LIMIT 1",
            (project_id,)
        )
        row = cur.fetchone()
    if not row:
        return jsonify({'error': 'Nenhuma análise em andamento'}), 400

    analysis_id = row['id']
    _cancel_set.add(analysis_id)

    # Update status immediately so polling detects it right away
    with db.cursor() as cur:
        cur.execute(
            '''UPDATE analyses SET status=%s, completed_at=%s, error_message=%s
               WHERE id=%s''',
            ('stopped', datetime.utcnow(), 'Interrompida pelo usuário', analysis_id)
        )
        db.commit()

    return jsonify({'ok': True, 'analysis_id': analysis_id})


@app.route('/admin/projetos/<int:project_id>/importar-sf', methods=['POST'])
@admin_required
def admin_sf_upload(project_id):
    """Receive uploaded Screaming Frog CSV files, create analysis and process in background."""
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT * FROM projects WHERE id = %s', (project_id,))
        project = cur.fetchone()
        if not project:
            return jsonify({'error': 'Projeto não encontrado'}), 404

        cur.execute(
            "SELECT id FROM analyses WHERE project_id=%s AND status IN ('pending','running')",
            (project_id,)
        )
        if cur.fetchone():
            return jsonify({'error': 'Já existe uma análise em andamento para este projeto'}), 400

    internal_file = request.files.get('internal_csv')
    images_file   = request.files.get('images_csv')

    if not internal_file or internal_file.filename == '':
        return jsonify({'error': 'Arquivo "Internal:All" é obrigatório'}), 400

    internal_bytes = internal_file.read()
    images_bytes   = images_file.read() if images_file and images_file.filename else None

    # Parse now (in request context) — fast, just CSV reading
    try:
        sf_results = process_sf_csv(internal_bytes, images_bytes)
    except Exception as e:
        return jsonify({'error': f'Erro ao ler o CSV: {e}'}), 400

    if not sf_results:
        return jsonify({'error': 'Nenhuma página HTML encontrada no CSV. Verifique o arquivo.'}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            "INSERT INTO analyses (project_id, status) VALUES (%s, 'pending') RETURNING id",
            (project_id,)
        )
        analysis_id = cur.fetchone()['id']
        db.commit()

    thread = threading.Thread(
        target=run_sf_import_background,
        args=(analysis_id, sf_results),
        daemon=True
    )
    thread.start()

    return jsonify({'analysis_id': analysis_id, 'status': 'started', 'pages': len(sf_results)})


@app.route('/admin/projetos/<int:project_id>/sf-cli', methods=['POST'])
@admin_required
def admin_sf_cli(project_id):
    """Trigger Screaming Frog CLI crawl for this project."""
    if not sf_cli_available():
        return jsonify({'error': 'Screaming Frog CLI não encontrado. Configure SF_CLI_PATH no .env'}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT * FROM projects WHERE id = %s', (project_id,))
        project = cur.fetchone()
        if not project:
            return jsonify({'error': 'Projeto não encontrado'}), 404

        cur.execute(
            "SELECT id FROM analyses WHERE project_id=%s AND status IN ('pending','running')",
            (project_id,)
        )
        if cur.fetchone():
            return jsonify({'error': 'Já existe uma análise em andamento para este projeto'}), 400

        cur.execute(
            "INSERT INTO analyses (project_id, status) VALUES (%s, 'pending') RETURNING id",
            (project_id,)
        )
        analysis_id = cur.fetchone()['id']
        db.commit()

    thread = threading.Thread(
        target=run_sf_cli_background,
        args=(analysis_id, project['url']),
        daemon=True
    )
    thread.start()

    return jsonify({'analysis_id': analysis_id, 'status': 'started'})


@app.route('/api/sf-cli-available')
@admin_required
def api_sf_cli_available():
    return jsonify({'available': sf_cli_available()})


@app.route('/admin/projetos/<int:project_id>/excluir', methods=['POST'])
@admin_required
def admin_delete_project(project_id):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('DELETE FROM projects WHERE id = %s', (project_id,))
        db.commit()
    flash('Projeto excluído com sucesso.', 'success')
    return redirect(url_for('admin_projects'))


@app.route('/admin/analise/<int:analysis_id>/pagina/<int:page_id>')
@admin_required
def admin_page_issues(analysis_id, page_id):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT * FROM pages WHERE id = %s AND analysis_id = %s', (page_id, analysis_id))
        page = cur.fetchone()
        if not page:
            return jsonify({'error': 'Página não encontrada'}), 404

        cur.execute(
            'SELECT * FROM issues WHERE page_id = %s ORDER BY CASE severity WHEN %s THEN 1 WHEN %s THEN 2 ELSE 3 END',
            (page_id, 'high', 'medium')
        )
        issues = cur.fetchall()

    return jsonify({
        'page': dict(page),
        'issues': [dict(i) for i in issues]
    })


# ---------------------------------------------------------------------------
# API - Analysis status (polling)
# ---------------------------------------------------------------------------

@app.route('/api/analysis/<int:analysis_id>/status')
@login_required
def api_analysis_status(analysis_id):
    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT * FROM analyses WHERE id = %s', (analysis_id,))
        analysis = cur.fetchone()
        if not analysis:
            return jsonify({'error': 'Não encontrado'}), 404

    result = dict(analysis)
    # Ensure datetime fields are serializable
    for k in ('started_at', 'completed_at', 'created_at'):
        if result.get(k):
            result[k] = result[k].isoformat()
    return jsonify(result)


@app.route('/api/analysis/<int:analysis_id>/pages')
@login_required
def api_analysis_pages(analysis_id):
    """Return pages already analyzed so far (used for live preview during running analysis)."""
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            'SELECT id FROM analyses WHERE id = %s',
            (analysis_id,)
        )
        if not cur.fetchone():
            return jsonify({'error': 'Não encontrado'}), 404

        cur.execute('''
            SELECT pg.id, pg.url, pg.title, pg.status_code, pg.load_time,
                   COUNT(i.id) as total_issue_count,
                   COUNT(CASE WHEN i.severity='high' THEN 1 END) as high_count,
                   COUNT(CASE WHEN i.severity='medium' THEN 1 END) as medium_count,
                   COUNT(CASE WHEN i.severity='low' THEN 1 END) as low_count
            FROM pages pg
            LEFT JOIN issues i ON i.page_id = pg.id
            WHERE pg.analysis_id = %s
            GROUP BY pg.id
            ORDER BY pg.analyzed_at DESC
        ''', (analysis_id,))
        pages = [dict(r) for r in cur.fetchall()]

    return jsonify({'pages': pages})


# ---------------------------------------------------------------------------
# Client routes
# ---------------------------------------------------------------------------

@app.route('/cliente')
@login_required
def client_dashboard():
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))

    db = get_db()
    with db.cursor() as cur:
        cur.execute('''
            SELECT p.*,
                   (SELECT status FROM analyses WHERE project_id=p.id ORDER BY created_at DESC LIMIT 1) as last_status,
                   (SELECT total_issues FROM analyses WHERE project_id=p.id AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1) as total_issues,
                   (SELECT high_issues FROM analyses WHERE project_id=p.id AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1) as high_issues,
                   (SELECT completed_at FROM analyses WHERE project_id=p.id AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1) as last_analysis
            FROM projects p
            WHERE p.client_id = %s
            ORDER BY p.created_at DESC
        ''', (session['user_id'],))
        projects = cur.fetchall()

    return render_template('client/dashboard.html', projects=projects)


@app.route('/cliente/projeto/<int:project_id>')
@login_required
def client_project_view(project_id):
    if session.get('role') == 'admin':
        return redirect(url_for('admin_project_detail', project_id=project_id))

    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            'SELECT * FROM projects WHERE id = %s AND client_id = %s',
            (project_id, session['user_id'])
        )
        project = cur.fetchone()

        if not project:
            flash('Projeto não encontrado.', 'danger')
            return redirect(url_for('client_dashboard'))

        cur.execute(
            "SELECT * FROM analyses WHERE project_id=%s AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1",
            (project_id,)
        )
        latest_analysis = cur.fetchone()

        pages_data = []
        if latest_analysis:
            cur.execute('''
                SELECT pg.*,
                       COUNT(i.id) as total_issue_count,
                       COUNT(CASE WHEN i.severity='high' THEN 1 END) as high_count,
                       COUNT(CASE WHEN i.severity='medium' THEN 1 END) as medium_count,
                       COUNT(CASE WHEN i.severity='low' THEN 1 END) as low_count,
                       COALESCE(array_agg(DISTINCT i.category) FILTER (WHERE i.category IS NOT NULL), '{}') as issue_categories,
                       COALESCE(array_agg(DISTINCT i.severity) FILTER (WHERE i.severity IS NOT NULL), '{}') as issue_severities
                FROM pages pg
                LEFT JOIN issues i ON i.page_id = pg.id
                WHERE pg.analysis_id = %s
                GROUP BY pg.id
                ORDER BY COUNT(i.id) DESC
            ''', (latest_analysis['id'],))
            pages_data = cur.fetchall()

    return render_template('client/project_view.html',
                           project=project,
                           latest_analysis=latest_analysis,
                           pages_data=pages_data)


@app.route('/cliente/projeto/<int:project_id>/pagina/<int:page_id>')
@login_required
def client_page_issues(project_id, page_id):
    db = get_db()
    with db.cursor() as cur:
        # Validate ownership
        cur.execute('SELECT id FROM projects WHERE id=%s AND client_id=%s', (project_id, session['user_id']))
        if not cur.fetchone() and session.get('role') != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403

        cur.execute('SELECT * FROM pages WHERE id = %s', (page_id,))
        page = cur.fetchone()

        cur.execute(
            'SELECT * FROM issues WHERE page_id = %s ORDER BY CASE severity WHEN %s THEN 1 WHEN %s THEN 2 ELSE 3 END',
            (page_id, 'high', 'medium')
        )
        issues = cur.fetchall()

    return jsonify({
        'page': dict(page),
        'issues': [dict(i) for i in issues]
    })


# ---------------------------------------------------------------------------
# Validate corrections – re-crawl a single page and compare issues
# ---------------------------------------------------------------------------

@app.route('/cliente/projeto/<int:project_id>/validar/<int:page_id>', methods=['POST'])
@login_required
def client_validate_page(project_id, page_id):
    from crawler import fetch_page
    from analyzer import analyze_page

    db = get_db()
    with db.cursor() as cur:
        # Validate ownership
        if session.get('role') == 'admin':
            cur.execute('SELECT id FROM projects WHERE id=%s', (project_id,))
        else:
            cur.execute('SELECT id FROM projects WHERE id=%s AND client_id=%s',
                        (project_id, session['user_id']))
        if not cur.fetchone():
            return jsonify({'error': 'Acesso negado'}), 403

        # Get existing page and its issues
        cur.execute('SELECT * FROM pages WHERE id=%s', (page_id,))
        page = cur.fetchone()
        if not page:
            return jsonify({'error': 'Página não encontrada'}), 404

        cur.execute(
            'SELECT * FROM issues WHERE page_id=%s ORDER BY CASE severity WHEN %s THEN 1 WHEN %s THEN 2 ELSE 3 END',
            (page_id, 'high', 'medium')
        )
        old_issues = cur.fetchall()

    # Build set of old issue identifiers (category + title)
    old_issue_keys = {(i['category'], i['title']) for i in old_issues}

    # Re-crawl the URL
    url = page['url']
    page_data = fetch_page(url)
    if not page_data:
        return jsonify({'error': 'Não foi possível acessar a URL'}), 500

    # Re-analyze
    new_issues_raw, word_count = analyze_page(page_data)
    new_issue_keys = {(i['category'], i['title']) for i in new_issues_raw}

    # Compare
    fixed = []
    for iss in old_issues:
        key = (iss['category'], iss['title'])
        if key not in new_issue_keys:
            fixed.append(dict(iss))

    remaining = []
    for iss in new_issues_raw:
        key = (iss['category'], iss['title'])
        if key in old_issue_keys:
            remaining.append(iss)

    new_found = []
    for iss in new_issues_raw:
        key = (iss['category'], iss['title'])
        if key not in old_issue_keys:
            new_found.append(iss)

    return jsonify({
        'url': url,
        'title': page_data.get('title', ''),
        'old_count': len(old_issues),
        'fixed': fixed,
        'remaining': remaining,
        'new_found': new_found,
        'fixed_count': len(fixed),
        'remaining_count': len(remaining),
        'new_count': len(new_found),
    })


# ---------------------------------------------------------------------------
# Export issues to Excel spreadsheet
# ---------------------------------------------------------------------------

@app.route('/cliente/projeto/<int:project_id>/exportar')
@login_required
def client_export_issues(project_id):
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    with db.cursor() as cur:
        # Validate ownership
        if session.get('role') == 'admin':
            cur.execute('SELECT * FROM projects WHERE id=%s', (project_id,))
        else:
            cur.execute('SELECT * FROM projects WHERE id=%s AND client_id=%s',
                        (project_id, session['user_id']))
        project = cur.fetchone()
        if not project:
            flash('Projeto não encontrado.', 'danger')
            return redirect(url_for('client_dashboard'))

        cur.execute(
            "SELECT * FROM analyses WHERE project_id=%s AND status IN ('completed','stopped') ORDER BY completed_at DESC LIMIT 1",
            (project_id,)
        )
        analysis = cur.fetchone()
        if not analysis:
            flash('Nenhuma análise disponível para exportar.', 'warning')
            return redirect(url_for('client_project_view', project_id=project_id))

        cur.execute('''
            SELECT pg.url AS page_url, pg.title AS page_title,
                   i.category, i.severity, i.title, i.description,
                   i.current_value, i.suggestion, i.ai_suggestion
            FROM pages pg
            JOIN issues i ON i.page_id = pg.id
            WHERE pg.analysis_id = %s
            ORDER BY pg.url,
                     CASE i.severity WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END
        ''', (analysis['id'],))
        rows = cur.fetchall()

    # -- Build Excel workbook --
    wb = Workbook()
    ws = wb.active
    ws.title = 'Problemas SEO'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    sev_fills = {
        'high': PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid'),
        'medium': PatternFill(start_color='FFFBF0', end_color='FFFBF0', fill_type='solid'),
        'low': PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid'),
    }
    sev_fonts = {
        'high': Font(bold=True, color='DC3545'),
        'medium': Font(bold=True, color='B45309'),
        'low': Font(color='6C757D'),
    }
    sev_labels = {'high': 'Crítico', 'medium': 'Médio', 'low': 'Baixo'}
    cat_labels = {
        'title': 'Título', 'meta': 'Meta Description', 'heading': 'Títulos (H1/H2)',
        'content': 'Conteúdo', 'image': 'Imagens', 'social': 'Redes Sociais',
        'technical': 'Técnico', 'links': 'Links', 'legibilidade': 'Legibilidade'
    }

    headers = ['URL da Página', 'Título da Página', 'Categoria', 'Severidade',
               'Problema', 'Descrição', 'Situação Atual', 'Sugestão', 'Sugestão IA']

    # Write header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        # Parse AI suggestion
        ai_text = ''
        if row['ai_suggestion']:
            try:
                ai = row['ai_suggestion'] if isinstance(row['ai_suggestion'], dict) else json.loads(row['ai_suggestion'])
                parts = []
                if ai.get('titulo'):
                    parts.append(f"Título: {ai['titulo']}")
                if ai.get('meta_description'):
                    parts.append(f"Meta: {ai['meta_description']}")
                if ai.get('h1'):
                    parts.append(f"H1: {ai['h1']}")
                if ai.get('dica_conteudo'):
                    parts.append(f"Dica: {ai['dica_conteudo']}")
                if ai.get('palavras_chave'):
                    parts.append(f"Palavras-chave: {', '.join(ai['palavras_chave'])}")
                ai_text = '\n'.join(parts)
            except (json.JSONDecodeError, TypeError):
                ai_text = str(row['ai_suggestion'])

        sev = row['severity']
        values = [
            row['page_url'],
            row['page_title'] or '(sem título)',
            cat_labels.get(row['category'], row['category']),
            sev_labels.get(sev, sev),
            row['title'],
            row['description'] or '',
            row['current_value'] or '',
            row['suggestion'] or '',
            ai_text
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Color the severity cell
        sev_cell = ws.cell(row=row_idx, column=4)
        sev_cell.fill = sev_fills.get(sev, PatternFill())
        sev_cell.font = sev_fonts.get(sev, Font())

    # Column widths
    widths = [45, 35, 18, 12, 30, 40, 30, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = f'A1:I{len(rows) + 1}'

    # -- Summary sheet --
    ws2 = wb.create_sheet('Resumo')
    ws2.cell(row=1, column=1, value='Relatório SEO').font = Font(bold=True, size=14, color='2563EB')
    ws2.cell(row=2, column=1, value=project['name']).font = Font(bold=True, size=12)
    ws2.cell(row=3, column=1, value=project['url'])
    ws2.cell(row=4, column=1, value=f"Data da análise: {analysis['completed_at'].strftime('%d/%m/%Y %H:%M') if analysis.get('completed_at') else ''}")
    ws2.cell(row=6, column=1, value='Resumo').font = Font(bold=True, size=11)
    ws2.cell(row=7, column=1, value='Total de páginas:')
    ws2.cell(row=7, column=2, value=analysis['total_pages'])
    ws2.cell(row=8, column=1, value='Problemas críticos:')
    ws2.cell(row=8, column=2, value=analysis['high_issues']).font = Font(bold=True, color='DC3545')
    ws2.cell(row=9, column=1, value='Problemas médios:')
    ws2.cell(row=9, column=2, value=analysis['medium_issues']).font = Font(bold=True, color='B45309')
    ws2.cell(row=10, column=1, value='Problemas baixos:')
    ws2.cell(row=10, column=2, value=analysis['low_issues'])
    ws2.cell(row=11, column=1, value='Total de problemas:')
    ws2.cell(row=11, column=2, value=analysis['total_issues']).font = Font(bold=True)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio-seo-{project['name'].replace(' ', '-').lower()}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ---------------------------------------------------------------------------
# On-demand AI suggestion for title / meta description
# ---------------------------------------------------------------------------

@app.route('/api/page/<int:page_id>/gerar-sugestao', methods=['POST'])
@login_required
def api_generate_suggestion(page_id):
    tipo = (request.json or {}).get('tipo', 'meta')  # 'titulo' or 'meta'

    db = get_db()
    with db.cursor() as cur:
        cur.execute('SELECT url, title FROM pages WHERE id = %s', (page_id,))
        page = cur.fetchone()
        if not page:
            return jsonify({'error': 'Página não encontrada'}), 404

    api_key = os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'Chave Gemini não configurada'}), 500

    try:
        import requests as req_lib
        import google.generativeai as genai
        from bs4 import BeautifulSoup as BS

        resp = req_lib.get(page['url'], timeout=10,
                           headers={'User-Agent': 'Mozilla/5.0 (compatible; SEOBot/1.0)'})
        soup = BS(resp.text, 'html.parser')

        h1_tags = soup.find_all('h1')
        h1_text = h1_tags[0].get_text().strip() if h1_tags else ''
        h2_texts = [h.get_text().strip() for h in soup.find_all('h2')[:5]]
        body = soup.find('body')
        body_text = body.get_text(separator=' ', strip=True)[:2000] if body else ''

        if tipo == 'titulo':
            instrucao = 'Gere APENAS um título de página (title tag) otimizado para SEO com 50-60 caracteres.'
            campo = 'titulo'
        else:
            instrucao = 'Gere APENAS uma meta description otimizada para SEO com 150-160 caracteres, incluindo call-to-action.'
            campo = 'meta_description'

        prompt = f"""Você é um especialista em SEO e copywriting. {instrucao}

URL: {page['url']}
Título atual: {page['title'] or '(ausente)'}
H1: {h1_text or '(ausente)'}
Subtítulos H2: {', '.join(h2_texts) if h2_texts else '(nenhum)'}
Conteúdo da página: {body_text}

Retorne APENAS um JSON válido sem markdown: {{"{campo}": "texto gerado aqui"}}
Use português brasileiro. Seja específico ao tema da página."""

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        response = model.generate_content(prompt)
        response_text = response.text.strip()

        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            data = json.loads(json_match.group())
            return jsonify({'sugestao': data.get(campo, ''), 'tipo': tipo})

    except Exception as e:
        print(f'[AI Suggest] Erro: {e}')
        return jsonify({'error': str(e)}), 500

    return jsonify({'error': 'Não foi possível gerar sugestão'}), 500


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    print('[INIT] Inicializando banco de dados...')
    init_db()
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=debug)
